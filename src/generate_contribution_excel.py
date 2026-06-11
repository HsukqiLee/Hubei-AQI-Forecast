import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_contribution_table():
    print("正在生成 小组分工与贡献表.xlsx...")
    
    # 1. 创建 DataFrame
    data = [
        {
            "姓名": "[组长姓名]",
            "角色": "组长",
            "学号": "[组长学号]",
            "主要负责模块": "机器学习建模 (model_training.py)",
            "具体工作描述": "项目总体架构规划与任务整合；实现LSTM深度学习模型构建、滑动窗口数据集划分、模型训练及参数调优；编写核心集成流水线 run_pipeline.py；负责报告中模型与模型拟合部分撰写；课堂汇报核心技术答辩。",
            "估计代码行数": 150,
            "贡献度占比": "28%"
        },
        {
            "姓名": "[组员1姓名]",
            "角色": "组员",
            "学号": "[组员1学号]",
            "主要负责模块": "数据采集与清洗 (fetch_real_2026_data.py, data_preprocessing.py)",
            "具体工作描述": "编写脚本对接 Open-Meteo API 抓取湖北省 17 个地级市真实空气质量监测数据；实现 17 地市 Excel 数据逆序重整与大表合并，输出 AirCondition.csv；负责数据采集部分的文档撰写与课堂 PPT 第一章节汇报。",
            "估计代码行数": 120,
            "贡献度占比": "24%"
        },
        {
            "姓名": "[组员2姓名]",
            "角色": "组员",
            "学号": "[组员2学号]",
            "主要负责模块": "数据探索与特征构造 (data_exploration.py, feature_engineering.py)",
            "具体工作描述": "实现AQI在17地市的均值分析、多地市走势折线图绘制、以及7种污染物相关性热力图生成；编写归一化Scaler、1-7天时间滞后特征构造、SelectKBest特征工程筛选脚本；负责报告中数据分析章节撰写与课堂PPT特征分析章节讲解。",
            "估计代码行数": 100,
            "贡献度占比": "24%"
        },
        {
            "姓名": "[组员3姓名]",
            "角色": "组员",
            "学号": "[组员3学号]",
            "主要负责模块": "数据可视化与报告填写 (visualization.py, fill_report.py)",
            "具体工作描述": "实现17地市历史与预测融合对比子图绘制（修复案例原版索引溢出Bug）；基于pyecharts生成Hubei交互式空气预测地图网页；负责整理撰写实验报告书中的可视化图表与文字内容，输出完整报告；负责PPT排版及成果展示环节讲解。",
            "估计代码行数": 130,
            "贡献度占比": "24%"
        }
    ]
    
    df = pd.DataFrame(data)
    
    # 使用 pandas/openpyxl 保存
    import os
    os.makedirs("reports", exist_ok=True)
    file_path = "reports/小组分工与贡献表.xlsx"
    df.to_excel(file_path, index=False)
    
    # 使用 openpyxl 设计 Excel 文件样式以获得高级外观
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.title = "分工与贡献度"
    
    # 将工作表视图网格线设置为 True
    ws.views.sheetView[0].showGridLines = True
    
    # 字体和对齐方式
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Microsoft YaHei", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # 深蓝色
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # 柔和的蓝灰色
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom = Side(border_style="medium", color="1F4E78")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 设置标题行样式
    ws.row_dimensions[1].height = 28
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
        
    # 设置数据行样式
    for row_idx in range(2, 6):
        ws.row_dimensions[row_idx].height = 70  # 高行以处理换行文本
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = cell_border
            
            # 对齐方式
            if col_idx in [1, 2, 3, 6, 7]: # 姓名, 角色, 学号, 代码行数, 贡献
                cell.alignment = center_align
            else: # 模块, 工作描述
                cell.alignment = left_align
                
            # 斑马条纹
            if is_even:
                cell.fill = zebra_fill
                
    # 列宽
    ws.column_dimensions['A'].width = 12  # 姓名
    ws.column_dimensions['B'].width = 10  # 角色
    ws.column_dimensions['C'].width = 16  # 学号
    ws.column_dimensions['D'].width = 30  # 模块
    ws.column_dimensions['E'].width = 60  # 工作描述
    ws.column_dimensions['F'].width = 14  # 代码行数
    ws.column_dimensions['G'].width = 12  # 贡献
    
    wb.save(file_path)
    print("成功设置样式并保存为 reports/小组分工与贡献表.xlsx")

if __name__ == "__main__":
    generate_contribution_table()
